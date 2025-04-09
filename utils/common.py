import arcpy
import openpyxl
import os
import sys
from pathlib import Path
from importlib import reload


ROOT = str(Path(__file__).parents[1].absolute()) # ../pyt_framework

if ROOT not in sys.path:
    sys.path.insert(0, ROOT)
if rf"{ROOT}\enums" not in sys.path:
    sys.path.insert(1, rf"{ROOT}\enums")

# Inline reloader of dynamic modules
[
    print(f"Reloaded {reload(module).__name__}")
    for module_name, module in globals().items()
    if module_name.startswith("pyt_reload")
]

# Import the Tool Importer function
import enums.domains as domains
from enums.excel_values import ExcelColumns


def load_excel_data(excel_path, sheet_name):

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    return ws

def get_column_indices(ws):
    header = [cell.value for cell in ws[1]]
    return {col: header.index(col) for col in header}

def get_row_values(row, col_idx):
    return {
        ExcelColumns.NAME.value: row[col_idx[ExcelColumns.NAME.value]],
        ExcelColumns.ALIAS.value: row[col_idx[ExcelColumns.ALIAS.value]],
        ExcelColumns.TYPE.value: row[col_idx[ExcelColumns.TYPE.value]],
        ExcelColumns.DOMAIN.value: row[col_idx[ExcelColumns.DOMAIN.value]],
        ExcelColumns.DEFAULT_VALUE.value: row[col_idx[ExcelColumns.DEFAULT_VALUE.value]]
    }

def add_fields_from_excel_to_layer(ws, layer_name, gdb_path):
    col_idx = get_column_indices(ws)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx[ExcelColumns.TABLE_NAME.value]] == layer_name:
            field_params = get_row_values(row, col_idx)
            create_field(gdb_path, layer_name, **field_params)

def create_field(output_gdb, layer_name, **field_params):
    output_path = os.path.join(output_gdb, layer_name)

    existing_fields = [field.name for field in arcpy.ListFields(output_path)]
    if field_params[ExcelColumns.NAME.value] not in existing_fields:
        arcpy.AddField_management(output_path, field_params[ExcelColumns.NAME.value],
                                  field_params[ExcelColumns.TYPE.value], field_alias=field_params[ExcelColumns.ALIAS.value])

    if field_params[ExcelColumns.DOMAIN.value]:
        existing_domains = [d.name for d in arcpy.da.ListDomains(output_gdb)]
        if field_params[ExcelColumns.DOMAIN.value] not in existing_domains:
            arcpy.CreateDomain_management(output_gdb, field_params[ExcelColumns.DOMAIN.value], field_type = "TEXT")

        domain_dict = {e.value: e.name for e in getattr(domains, field_params[ExcelColumns.DOMAIN.value])}
        for code in domain_dict:
            arcpy.AddCodedValueToDomain_management(output_gdb, field_params[ExcelColumns.DOMAIN.value], domain_dict[code], code)
        arcpy.AssignDomainToField_management(output_path, field_params[ExcelColumns.NAME.value], field_params[ExcelColumns.DOMAIN.value])

    if field_params[ExcelColumns.DEFAULT_VALUE.value]:
        arcpy.management.CalculateField(output_path, field_params[ExcelColumns.NAME.value],

                                        field_params[ExcelColumns.DEFAULT_VALUE.value], "PYTHON3")
